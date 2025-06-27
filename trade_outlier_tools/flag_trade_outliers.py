import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

input_base_dir = r'R:\Project 1\Output Trades (Tableau)'
output_dir = r'R:\Project 1\Output Calculations'


trade_date = input("Enter the trade date in format DDMMMYY: ").upper()

files = {
    'POWER': {
        'input': os.path.join(input_base_dir, f'output_power_trades_{trade_date}.csv'),
        'output': os.path.join(output_dir, f'POWER_{trade_date}_Zscores.xlsx')
    },
    'GAS': {
        'input': os.path.join(input_base_dir, f'output_gas_trades_{trade_date}.csv'),
        'output': os.path.join(output_dir, f'GAS_{trade_date}_Zscores.xlsx')
    }
}

def process_trade_file(input_csv_path, output_excel_path, label):
    print(f"\n Processing {label} data...")

    try:
        df = pd.read_csv(input_csv_path)
    except FileNotFoundError:
        print(f" File not found for {label}: {input_csv_path}")
        return

    if 'EDFT_Diff' in df.columns and 'TradeID' in df.columns:
        mean_diff = df['EDFT_Diff'].mean()
        std_diff = df['EDFT_Diff'].std(ddof=0)

        if std_diff != 0:
            df['Zscore'] = (df['EDFT_Diff'] - mean_diff) / std_diff
        else:
            df['Zscore'] = 0.0

        threshold = 2.5
        df['Flagged'] = df['Zscore'].abs() > threshold

        output_df = df[['TradeID', 'EDFT_Diff', 'Zscore', 'Flagged']]
        output_df.to_excel(output_excel_path, index=False)

        wb = load_workbook(output_excel_path)
        ws = wb.active
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_font = Font(color="006100", bold=True)

        for row in range(2, ws.max_row + 1):
            is_flagged = ws[f'D{row}'].value
            if is_flagged:
                for col in ['A', 'B', 'C', 'D']:
                    ws[f'{col}{row}'].fill = red_fill
                ws[f'D{row}'].font = green_font

        wb.save(output_excel_path)
        print(f" Z-score results saved and highlighted at:\n{output_excel_path}")
    else:
        print(f" Required columns missing in {label} input file: {input_csv_path}")

process_trade_file(files['POWER']['input'], files['POWER']['output'], 'POWER')
process_trade_file(files['GAS']['input'], files['GAS']['output'], 'GAS')
