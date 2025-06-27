import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from statsmodels.robust.scale import mad
from sklearn.ensemble import IsolationForest

# User input 
trade_date = input("Enter the trade date in format DDMMMYY (e.g., 25JUN25): ").upper()

# Input data (automate it with "Trade_Check1.py" code)
input_base_dir = r'R:\Project 1\Output Trades (Tableau)'
output_dir = r'R:\Project 1\Output Calculations'

files = {
    'POWER': {
        'input': os.path.join(input_base_dir, f'output_power_trades_{trade_date}.csv'),
        'output': os.path.join(output_dir, f'POWER_{trade_date}_Zscores.xlsx')
    },
    'GAS': {
        'input': os.path.join(input_base_dir, f'output_gas_trades_{trade_date}.csv'),
        'output': os.path.join(output_dir, f'GAS_{trade_date}_Zscores.xlsx')
    }
}

def process_trade_file(input_csv_path, output_excel_path, label):
    print(f"\nProcessing {label} data...")

    try:
        df = pd.read_csv(input_csv_path)
    except FileNotFoundError:
        print(f" File not found for {label}: {input_csv_path}")
        return

    if 'EDFT_Diff' in df.columns and 'TradeID' in df.columns:

        # === Z-score Method ===
        mean_diff = df['EDFT_Diff'].mean()
        std_diff = df['EDFT_Diff'].std(ddof=0)
        df['Zscore'] = (df['EDFT_Diff'] - mean_diff) / std_diff if std_diff != 0 else 0.0
        df['Flag_Zscore'] = df['Zscore'].abs() > 2.5

        # === Median + MAD Method ===
        median_diff = df['EDFT_Diff'].median()
        mad_diff = mad(df['EDFT_Diff'])
        if mad_diff != 0:
            df['RobustZ'] = (df['EDFT_Diff'] - median_diff) / (1.4826 * mad_diff)
        else:
            df['RobustZ'] = np.nan
        df['Flag_RobustZ'] = df['RobustZ'].abs() > 2.5

        # === Isolation Forest Method ===
        valid_iforest = df['EDFT_Diff'].replace([np.inf, -np.inf], np.nan).dropna()
        #valid_iforest = df[['EDFT_Diff', 'Volume']].replace([np.inf, -np.inf], np.nan).dropna()
        if not valid_iforest.empty:
            if_model = IsolationForest(contamination=0.05, random_state=42)
            preds = pd.Series(index=valid_iforest.index, data=if_model.fit_predict(valid_iforest.values.reshape(-1, 1)))
            df['IF_Score'] = preds
            df['Flag_IForest'] = df['IF_Score'] == -1
        else:
            df['IF_Score'] = np.nan
            df['Flag_IForest'] = False

        # === Combine Method Flags ===
        df['FlaggedBy'] = df.apply(lambda row: ', '.join([
            method for method, flag in zip(
                ['Zscore', 'RobustZ', 'IForest'],
                [row['Flag_Zscore'], row['Flag_RobustZ'], row['Flag_IForest']]
            ) if flag
        ]), axis=1)

        df['Flag_Count'] = df[['Flag_Zscore', 'Flag_RobustZ', 'Flag_IForest']].sum(axis=1)

        # === Write to Excel ===
        output_df = df[['TradeID', 'EDFT_Diff', 'Zscore', 'Flag_Zscore',
                        'RobustZ', 'Flag_RobustZ', 'Flag_IForest', 'FlaggedBy', 'Flag_Count']]
        output_df.to_excel(output_excel_path, index=False)

        # === Highlight rows flagged by ≥ 2 methods ===
        wb = load_workbook(output_excel_path)
        ws = wb.active
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        green_font = Font(color="006100", bold=True)

        for row in range(2, ws.max_row + 1):
            flag_count = ws[f'I{row}'].value 
            if flag_count >= 2:
                for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
                    ws[f'{col}{row}'].fill = red_fill
                ws[f'H{row}'].font = green_font 

        wb.save(output_excel_path)
        print(f" {label} results saved with flags at:\n{output_excel_path}")

    else:
        print(f" Required columns missing in {label} input file: {input_csv_path}")

process_trade_file(files['POWER']['input'], files['POWER']['output'], 'POWER')
process_trade_file(files['GAS']['input'], files['GAS']['output'], 'GAS')
